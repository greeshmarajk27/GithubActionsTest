
import difflib
import sys
import re
from pathlib import Path
import pandas as pd

# primary: strict DOM
import xml.dom.minidom

# secondary: tolerant parser (install lxml in workflow)
try:
    from lxml import etree
    HAVE_LXML = True
except Exception:
    HAVE_LXML = False

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

HTML_COLORS = {
    "Added": "aaffaa",    # light green
    "Changed": "ffff77",  # yellow
    "Deleted": "ffaaaa",  # light red
    "Equal": "ffffff"     # white
}

CTRL_CHARS_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

def load_text_safely(path: str) -> str:
    """
    Read file as bytes, decode with UTF-8 (strip BOM), remove control chars.
    This avoids common XML parsing crashes.
    """
    data = Path(path).read_bytes()
    # Remove UTF-8 BOM if present
    text = data.decode("utf-8-sig", errors="ignore")
    # Remove illegal XML control chars
    text = CTRL_CHARS_RE.sub("", text)
    return text

def pretty_xml_lines(path: str) -> list[str]:
    """
    Try strict pretty-print (minidom); if that fails, try lxml (recover=True);
    if that also fails, fall back to raw lines.
    Always return a list of non-empty, trimmed lines.
    """
    raw = load_text_safely(path).strip()

    if not raw:
        print(f"⚠️  File is empty or unreadable as text: {path}")
        return []

    # --- Attempt 1: strict DOM (minidom) ---
    try:
        dom = xml.dom.minidom.parseString(raw)
        pretty = dom.toprettyxml(indent="  ")
        lines = [ln for ln in pretty.splitlines() if ln.strip()]
        return lines
    except Exception as e1:
        print(f"ℹ️  minidom failed for {path}: {e1}")

    # --- Attempt 2: tolerant parse with lxml (recover=True) ---
    if HAVE_LXML:
        try:
            parser = etree.XMLParser(remove_blank_text=True, recover=True)
            root = etree.fromstring(raw.encode("utf-8"), parser=parser)
            pretty = etree.tostring(root, pretty_print=True, encoding="unicode")
            lines = [ln for ln in pretty.splitlines() if ln.strip()]
            return lines
        except Exception as e2:
            print(f"ℹ️  lxml(recover) failed for {path}: {e2}")

    # --- Fallback: raw text (still allow diff) ---
    print(f"🔁 Falling back to raw line diff for {path}")
    return [ln for ln in raw.splitlines() if ln.strip()]

def classify(tag: str) -> str:
    return {"insert": "Added", "delete": "Deleted", "replace": "Changed"}.get(tag, "Equal")

def generate_excel_diff(file1: str, file2: str, out_xlsx: str) -> None:
    xml1 = pretty_xml_lines(file1)
    xml2 = pretty_xml_lines(file2)

    # Quick diagnostics if something looks off
    if not xml1:
        print(f"⚠️  Warning: {file1} produced 0 lines after preprocessing.")
    if not xml2:
        print(f"⚠️  Warning: {file2} produced 0 lines after preprocessing.")

    matcher = difflib.SequenceMatcher(None, xml1, xml2)
    rows = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        left_block  = xml1[i1:i2]
        right_block = xml2[j1:j2]
        max_len = max(len(left_block), len(right_block))
        change_type = classify(tag)

        for k in range(max_len):
            lnum = i1 + k + 1 if k < len(left_block)  else ""
            ltxt = left_block[k]  if k < len(left_block)  else ""
            rnum = j1 + k + 1 if k < len(right_block) else ""
            rtxt = right_block[k] if k < len(right_block) else ""
            rows.append([lnum, ltxt, rnum, rtxt, change_type])

    df = pd.DataFrame(rows, columns=["File1 Line", "File1 Text", "File2 Line", "File2 Text", "Change"])

    out = Path(out_xlsx)
    out.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Diff", index=False)
        ws = writer.book["Diff"]

        # HTML-like colors + monospace font
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            change = row[4].value
            color = HTML_COLORS.get(change, "ffffff")
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                cell.font = Font(name="Consolas")

        # Column widths
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 70

    print(f"✅ Excel diff saved → {out_xlsx}")

def main():
    if len(sys.argv) != 4:
        print("Usage: python scripts/arxml_excel.py <file1.xml> <file2.xml> <output.xlsx>")
        sys.exit(1)
    f1, f2, out = sys.argv[1], sys.argv[2], sys.argv[3]
    generate_excel_diff(f1, f2, out)

if __name__ == "__main__":
    main()

    